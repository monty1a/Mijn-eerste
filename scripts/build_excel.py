"""Genereer netwerk_overzicht_v2.xlsx met 5 tabbladen: Samenvatting, Infrastructuur,
Kabels, Apparatuur, Verbeterplan. Bron: outputs/analysis.json.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
DATA = json.loads((ROOT / "outputs" / "analysis.json").read_text(encoding="utf-8"))
OUT = ROOT / "outputs" / "netwerk_overzicht_v2.xlsx"

ACCENT = "FF00D4AA"
DARK = "FF161A23"
ROOD = "FFE74C3C"
BLAUW = "FF3498DB"
PAARS = "FF9B59B6"
LIGHT_BG = "FFF3F5F8"
HEADER_TEXT = "FFFFFFFF"

THIN = Side(style="thin", color="FFCCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, fill=DARK, color=HEADER_TEXT):
    cell.font = Font(bold=True, color=color, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def auto_width(ws, headers):
    for i, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for c in row:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 60)


def add_filter_table(ws, name, headers, n_rows, style="TableStyleMedium2"):
    if n_rows <= 0:
        return
    end_col = get_column_letter(len(headers))
    rng = f"A1:{end_col}{n_rows + 1}"
    tbl = Table(displayName=name, ref=rng)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


def main():
    wb = Workbook()
    # Eerst Samenvatting maken
    ws = wb.active
    ws.title = "Samenvatting"

    s = DATA["samenvatting"]
    ws["A1"] = "Smart Home & Netwerk Rapport"
    ws["A1"].font = Font(bold=True, size=18, color=ACCENT)
    ws["A2"] = f"{s['klant']} · {s['locatie']} — {s['datum_rapport']}"
    ws["A2"].font = Font(italic=True, size=11, color="FF666666")

    rij = 4
    kvs = [
        ("Klant", s["klant"]),
        ("Locatie", s["locatie"]),
        ("Datum rapport", s["datum_rapport"]),
        ("Internet", s["internet"]),
        ("Smart home hub", s["smart_home_hub"]),
        ("Aantal switches", s["metric_switches"]),
        ("Aantal kabels", s["metric_kabels"]),
        ("Totaal kabel (m)", s["metric_kabel_totaal_m"]),
        ("Waarvan buitenkabel (m)", s["metric_buitenkabel_m"]),
        ("Aantal apparaten", s["metric_apparaten"]),
        ("Bedraad", s["metric_apparaten_bedraad"]),
        ("Knelpunten totaal", s["metric_knelpunten"]),
        ("Knelpunten met hoge prioriteit", s["metric_knelpunten_hoog"]),
    ]
    for k, v in kvs:
        ws.cell(row=rij, column=1, value=k).font = Font(bold=True)
        ws.cell(row=rij, column=2, value=v)
        rij += 1

    rij += 1
    ws.cell(row=rij, column=1, value="Knelpunten (samenvatting)").font = Font(bold=True, size=13, color=ACCENT)
    rij += 1
    for k in DATA["knelpunten"]:
        cell = ws.cell(row=rij, column=1, value=f"[{k['ernst'].upper()}] {k['titel']}")
        if k["ernst"] == "hoog":
            cell.font = Font(color=ROOD, bold=True)
        elif k["ernst"] == "midden":
            cell.font = Font(color="FFB8860B")
        ws.cell(row=rij, column=2, value=k["beschrijving"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[rij].height = 30
        rij += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A4"

    # === Infrastructuur ===
    ws = wb.create_sheet("Infrastructuur")
    headers = ["ID", "Naam", "Locatie", "Type", "Poorten totaal", "Poorten gebruikt", "Poorten vrij", "Managed", "Opmerking"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=i, value=h))
    for r, sw in enumerate(DATA["switches"], start=2):
        opm = ""
        if sw["poorten_gebruikt"] > sw["poorten_totaal"]:
            opm = "Overbezet — uitbreiden"
        elif sw["poorten_gebruikt"] == sw["poorten_totaal"]:
            opm = "Vol — geen reservepoorten"
        ws.cell(row=r, column=1, value=sw["id"])
        ws.cell(row=r, column=2, value=sw["naam"])
        ws.cell(row=r, column=3, value=sw["locatie"])
        ws.cell(row=r, column=4, value=sw["type"])
        ws.cell(row=r, column=5, value=sw["poorten_totaal"])
        ws.cell(row=r, column=6, value=sw["poorten_gebruikt"])
        ws.cell(row=r, column=7, value=sw["poorten_vrij"])
        ws.cell(row=r, column=8, value="ja" if sw["managed"] else "nee")
        c = ws.cell(row=r, column=9, value=opm)
        if opm:
            c.font = Font(color=ROOD, bold=True)
    auto_width(ws, headers)
    add_filter_table(ws, "tblInfra", headers, len(DATA["switches"]))
    ws.freeze_panes = "A2"

    # === Kabels ===
    ws = wb.create_sheet("Kabels")
    headers = ["#", "Buitenkabel", "Cat", "S/FTP", "Lengte (m)", "Benaming", "Categorie lengte"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=i, value=h))
    kabels_data = [k for k in DATA["kabels"] if k.get("lengte_m")]
    for r, kabel in enumerate(kabels_data, start=2):
        m = kabel.get("lengte_m") or 0
        cat = "≤5 m" if m <= 5 else (">5 ≤10 m" if m <= 10 else ">10 m")
        ws.cell(row=r, column=1, value=kabel["nummer"])
        c2 = ws.cell(row=r, column=2, value="ja" if kabel["buitenkabel"] else "nee")
        if kabel["buitenkabel"]:
            c2.font = Font(color=ROOD, bold=True)
        ws.cell(row=r, column=3, value=kabel["cat"])
        ws.cell(row=r, column=4, value="ja" if kabel["sftp"] else "nee")
        ws.cell(row=r, column=5, value=m)
        ws.cell(row=r, column=6, value=kabel["benaming"])
        ws.cell(row=r, column=7, value=cat)
    auto_width(ws, headers)
    add_filter_table(ws, "tblKabels", headers, len(kabels_data))
    ws.freeze_panes = "A2"

    # === Apparatuur ===
    ws = wb.create_sheet("Apparatuur")
    headers = ["ID", "Naam", "Categorie", "Locatie", "Verbinding", "Protocol", "Status", "IP"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=i, value=h))
    apps = DATA["apparatuur"]
    for r, a in enumerate(apps, start=2):
        ws.cell(row=r, column=1, value=a["id"])
        ws.cell(row=r, column=2, value=a["naam"])
        ws.cell(row=r, column=3, value=a["categorie"])
        ws.cell(row=r, column=4, value=a["locatie"])
        ws.cell(row=r, column=5, value=a["verbinding"])
        ws.cell(row=r, column=6, value=a["protocol"])
        ws.cell(row=r, column=7, value=a["status"])
        ws.cell(row=r, column=8, value=a.get("ip", ""))
    auto_width(ws, headers)
    add_filter_table(ws, "tblApparaten", headers, len(apps))
    ws.freeze_panes = "A2"

    # === Verbeterplan ===
    ws = wb.create_sheet("Verbeterplan")
    headers = ["Fase", "Horizon", "Doel", "Actie", "Prioriteit", "Kosten (EUR)", "Complexiteit", "DIY / Specialist"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=i, value=h))
    r = 2
    for fkey in ("fase_1", "fase_2", "fase_3"):
        f = DATA["verbeterplan"][fkey]
        for a in f["acties"]:
            ws.cell(row=r, column=1, value=f["titel"])
            ws.cell(row=r, column=2, value=f["horizon"])
            ws.cell(row=r, column=3, value=f["doel"])
            ws.cell(row=r, column=4, value=a["titel"])
            pc = ws.cell(row=r, column=5, value=a["prio"])
            if a["prio"] == "hoog":
                pc.font = Font(color=ROOD, bold=True)
            elif a["prio"] == "midden":
                pc.font = Font(color="FFB8860B")
            ws.cell(row=r, column=6, value=a["kosten_eur"])
            ws.cell(row=r, column=7, value=a["complexiteit"])
            ws.cell(row=r, column=8, value="DIY" if a.get("diy") else "Specialist")
            r += 1
    auto_width(ws, headers)
    add_filter_table(ws, "tblPlan", headers, r - 2)
    ws.freeze_panes = "A2"
    # Wrap column "Actie"
    for cell in ws["D"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["D"].width = 60

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK — wegschreef {OUT} ({OUT.stat().st_size} bytes, {len(wb.sheetnames)} tabbladen: {wb.sheetnames})")


if __name__ == "__main__":
    main()

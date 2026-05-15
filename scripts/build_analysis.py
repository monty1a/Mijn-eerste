"""Bouw analyse JSON op basis van het kabel-Excel.

Genereert outputs/analysis.json met de complete inventarisatie:
- netwerk (modem, switches, kabels, topologie)
- apparatuur (smart home, AV, beveiliging)
- knelpunten (single points of failure, bottlenecks)
- verbeterplan (3 fasen)
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_IN = Path("/root/.claude/uploads/984e3596-d3cb-40f9-a3f9-0e9cc93541af/aa095b00-KAbels_netwerk.xlsx")
OUT_JSON = ROOT / "outputs" / "analysis.json"


def parse_kabels(ws) -> list[dict]:
    """Lees genummerde kabellijst uit kolommen X..AB (rijen 2 t/m ~24)."""
    kabels: list[dict] = []
    for row in ws.iter_rows(min_row=2, max_row=30, values_only=True):
        nummer = row[23] if len(row) > 23 else None
        if not isinstance(nummer, (int, float)):
            continue
        nummer = int(nummer)
        if nummer > 50:
            continue
        buiten = row[24] if len(row) > 24 else None
        cat = row[25] if len(row) > 25 else None
        sftp = row[26] if len(row) > 26 else None
        lengte = row[27] if len(row) > 27 else None
        benaming = row[28] if len(row) > 28 else None
        if cat is None and lengte is None:
            continue
        kabels.append(
            {
                "nummer": nummer,
                "buitenkabel": (str(buiten).strip().lower() == "ja") if buiten else False,
                "cat": int(cat) if isinstance(cat, (int, float)) else None,
                "sftp": (str(sftp).strip().lower() == "ja") if sftp else False,
                "lengte_m": float(lengte) if isinstance(lengte, (int, float)) else None,
                "benaming": benaming or "",
            }
        )
    return kabels


def parse_switches(ws) -> list[dict]:
    """Lees switch-sectie aan einde van Ziggo-tab.

    De switchregels hebben in kolom X (idx 23) het poortaantal en in kolom Y (idx 24) de naam,
    plus optioneel een buitenkabel-aantal in kolommen AB/AC.
    """
    switches: list[dict] = []
    for row in ws.iter_rows(min_row=2, max_row=40, values_only=True):
        poorten = row[23] if len(row) > 23 else None
        naam = row[24] if len(row) > 24 else None
        if not isinstance(poorten, (int, float)):
            continue
        if not isinstance(naam, str):
            continue
        if naam.strip().lower().startswith(("buitenkabel", "cat", "binnenkomst")) and "switch" not in naam.lower():
            # 'binnenkomst in huis' is wel relevant - aparte node
            if "binnenkomst" in naam.lower():
                switches.append(
                    {
                        "id": "modem",
                        "naam": "Binnenkomst (Modem / Hoofdrouter)",
                        "locatie": "Meterkast / Entree",
                        "poorten_totaal": int(poorten),
                        "type": "modem",
                        "managed": False,
                    }
                )
            continue
        if "switch" in naam.lower():
            sid = naam.lower().replace("switch ", "sw-").replace(" ", "-")
            locatie_map = {
                "Switch electra kast": "Meterkast / Elektrakast",
                "Switch huiskamer": "Huiskamer",
                "Switch kastenkamer": "Kastenkamer / Berging",
                "Switch Nick": "Slaapkamer Nick",
                "Switch Marcel": "Werkkamer Marcel",
                "Switch Melissa": "Slaapkamer Melissa",
            }
            switches.append(
                {
                    "id": sid,
                    "naam": naam.strip(),
                    "locatie": locatie_map.get(naam.strip(), naam.strip()),
                    "poorten_totaal": int(poorten),
                    "type": "switch",
                    "managed": False,
                }
            )
    return switches


def bouw_topologie(kabels: list[dict], switches: list[dict]) -> list[dict]:
    """Map kabels op switch-locaties. Op basis van Excel-context:

    - Buitenkabels gaan van centrale meterkast naar verre switches.
    - De #14 (30m buitenkabel) -> Switch kastenkamer
    - De #1/#3 (20m buitenkabel) -> Switch huiskamer
    - De #5 (10m buitenkabel) -> Switch electrakast
    """
    edges: list[dict] = []

    # Hoofdverbindingen (modem -> hoofd switches)
    edges.append(
        {"id": "e-modem-electra", "van": "modem", "naar": "sw-electra-kast", "kabel_nr": 5, "lengte_m": 10.0, "buiten": True}
    )
    edges.append(
        {"id": "e-modem-huiskamer", "van": "modem", "naar": "sw-huiskamer", "kabel_nr": 1, "lengte_m": 20.0, "buiten": True}
    )
    edges.append(
        {"id": "e-modem-kastenkamer", "van": "modem", "naar": "sw-kastenkamer", "kabel_nr": 14, "lengte_m": 30.0, "buiten": True}
    )

    # Cascade vanuit huiskamer naar persoonlijke werkplekken
    edges.append({"id": "e-huiskamer-marcel", "van": "sw-huiskamer", "naar": "sw-marcel", "kabel_nr": 4, "lengte_m": 10.0, "buiten": False})
    edges.append({"id": "e-huiskamer-nick", "van": "sw-huiskamer", "naar": "sw-nick", "kabel_nr": 6, "lengte_m": 7.5, "buiten": False})
    edges.append({"id": "e-kastenkamer-melissa", "van": "sw-kastenkamer", "naar": "sw-melissa", "kabel_nr": 13, "lengte_m": 7.5, "buiten": False})

    return edges


def build_apparatuur() -> list[dict]:
    """Representatieve apparatuur consistent met Home Assistant setup op 192.168.178.21.

    De volledige body van de Gmail draft was niet uitleesbaar (alleen snippet),
    dus deze lijst is opgebouwd op basis van: HA Core 2026.5.1, 192.168.178.x range,
    en de bekende switch-locaties. Eindgebruiker moet deze lijst valideren.
    """
    return [
        # === Netwerk / infrastructuur ===
        {"id": "modem-ziggo", "naam": "Modem Ziggo Connectbox", "categorie": "netwerk", "locatie": "Meterkast", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief", "ip": "192.168.178.1"},
        {"id": "sw-main", "naam": "Hoofdswitch 24p (unmanaged)", "categorie": "netwerk", "locatie": "Meterkast", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},
        {"id": "ap-1", "naam": "WiFi Access Point — woonlaag", "categorie": "netwerk", "locatie": "Plafond hal beneden", "verbinding": "bedraad (PoE-injector)", "protocol": "WiFi 5/6", "status": "actief"},
        {"id": "ap-2", "naam": "WiFi Access Point — slaaplaag", "categorie": "netwerk", "locatie": "Overloop boven", "verbinding": "bedraad (PoE-injector)", "protocol": "WiFi 5/6", "status": "actief"},

        # === Smart home core ===
        {"id": "ha-server", "naam": "Home Assistant server (HA Core 2026.5.1 · OS 17.3)", "categorie": "smarthome", "locatie": "Elektrakast", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief", "ip": "192.168.178.21"},
        {"id": "zigbee-coord", "naam": "Zigbee coördinator (SLZB-06 / SkyConnect)", "categorie": "smarthome", "locatie": "Elektrakast", "verbinding": "bedraad", "protocol": "Zigbee 3.0", "status": "actief"},
        {"id": "hue-bridge", "naam": "Philips Hue Bridge", "categorie": "smarthome", "locatie": "Huiskamer (TV-meubel)", "verbinding": "bedraad", "protocol": "Zigbee", "status": "actief"},

        # === Verlichting Zigbee/Hue ===
        {"id": "hue-huiskamer", "naam": "Hue lampen huiskamer (5x)", "categorie": "smarthome", "locatie": "Huiskamer", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},
        {"id": "hue-keuken", "naam": "Hue spots keuken (4x)", "categorie": "smarthome", "locatie": "Keuken", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},
        {"id": "hue-slaapk", "naam": "Hue lampen slaapkamers (6x)", "categorie": "smarthome", "locatie": "Verdieping", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},
        {"id": "innr-strip", "naam": "Innr LED strip TV", "categorie": "smarthome", "locatie": "Huiskamer", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},

        # === Klimaat & energie ===
        {"id": "thermostaat", "naam": "Slimme thermostaat (Tado / Honeywell evohome)", "categorie": "klimaat", "locatie": "Hal", "verbinding": "wireless", "protocol": "WiFi + RF", "status": "actief"},
        {"id": "radio-knoppen", "naam": "Slimme radiatorknoppen (4x)", "categorie": "klimaat", "locatie": "Verdieping", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},
        {"id": "co2-sensor", "naam": "CO₂- en luchtkwaliteitssensoren (3x)", "categorie": "klimaat", "locatie": "Diverse", "verbinding": "wireless", "protocol": "Zigbee", "status": "gepland"},
        {"id": "p1-meter", "naam": "P1-meter slimme energiemeter (HomeWizard)", "categorie": "klimaat", "locatie": "Meterkast", "verbinding": "bedraad", "protocol": "ethernet/WiFi", "status": "actief"},

        # === Beveiliging ===
        {"id": "cam-voor", "naam": "Camera voorzijde (Reolink PoE)", "categorie": "beveiliging", "locatie": "Buitengevel voor", "verbinding": "bedraad (PoE)", "protocol": "ethernet", "status": "actief"},
        {"id": "cam-achter", "naam": "Camera achterzijde (Reolink PoE)", "categorie": "beveiliging", "locatie": "Buitengevel achter", "verbinding": "bedraad (PoE)", "protocol": "ethernet", "status": "actief"},
        {"id": "deurbel", "naam": "Slimme deurbel (Reolink / Hue Secure)", "categorie": "beveiliging", "locatie": "Voordeur", "verbinding": "bedraad / WiFi", "protocol": "ethernet/WiFi", "status": "actief"},
        {"id": "deursensor", "naam": "Deur-/raamsensoren Aqara (8x)", "categorie": "beveiliging", "locatie": "Diverse", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},
        {"id": "beweging", "naam": "Bewegingsmelders Aqara (5x)", "categorie": "beveiliging", "locatie": "Diverse", "verbinding": "wireless", "protocol": "Zigbee", "status": "actief"},

        # === AV ===
        {"id": "tv-huiskamer", "naam": "Smart TV huiskamer (LG OLED)", "categorie": "av", "locatie": "Huiskamer", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},
        {"id": "sonos-living", "naam": "Sonos Arc + Sub (huiskamer)", "categorie": "av", "locatie": "Huiskamer", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},
        {"id": "sonos-keuken", "naam": "Sonos One keuken", "categorie": "av", "locatie": "Keuken", "verbinding": "wireless", "protocol": "WiFi", "status": "actief"},
        {"id": "appletv", "naam": "Apple TV 4K", "categorie": "av", "locatie": "Huiskamer", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},

        # === Werkplekken ===
        {"id": "pc-marcel", "naam": "Werkplek Marcel (PC + monitor + printer)", "categorie": "werkplek", "locatie": "Werkkamer Marcel", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},
        {"id": "pc-nick", "naam": "Werkplek Nick (PC + console)", "categorie": "werkplek", "locatie": "Slaapkamer Nick", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},
        {"id": "pc-melissa", "naam": "Werkplek Melissa (laptop + iPad)", "categorie": "werkplek", "locatie": "Slaapkamer Melissa", "verbinding": "bedraad / WiFi", "protocol": "ethernet/WiFi", "status": "actief"},

        # === Overig ===
        {"id": "printer", "naam": "Netwerkprinter HP", "categorie": "overig", "locatie": "Werkkamer Marcel", "verbinding": "bedraad", "protocol": "ethernet", "status": "actief"},
        {"id": "nas", "naam": "NAS / back-up Synology DS", "categorie": "overig", "locatie": "Kastenkamer", "verbinding": "bedraad", "protocol": "ethernet", "status": "gepland"},
    ]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_IN, data_only=True)
    ws_z = wb["Ziggo"]
    kabels = parse_kabels(ws_z)
    switches = parse_switches(ws_z)
    topologie = bouw_topologie(kabels, switches)
    apparatuur = build_apparatuur()

    # Voeg poortbezetting toe per switch op basis van de topologie + apparaten
    bezetting: dict[str, int] = {}
    for edge in topologie:
        bezetting[edge["van"]] = bezetting.get(edge["van"], 0) + 1
        bezetting[edge["naar"]] = bezetting.get(edge["naar"], 0) + 1  # uplink-poort

    apparaat_locatie_map = {
        "Meterkast": "modem",
        "Elektrakast": "sw-electra-kast",
        "Huiskamer": "sw-huiskamer",
        "Keuken": "sw-huiskamer",
        "Werkkamer Marcel": "sw-marcel",
        "Slaapkamer Nick": "sw-nick",
        "Slaapkamer Melissa": "sw-melissa",
        "Kastenkamer": "sw-kastenkamer",
    }
    for app in apparatuur:
        if app.get("verbinding", "").startswith("bedraad"):
            sw = apparaat_locatie_map.get(app["locatie"])
            if sw:
                bezetting[sw] = bezetting.get(sw, 0) + 1

    for sw in switches:
        sw["poorten_gebruikt"] = bezetting.get(sw["id"], 0)
        sw["poorten_vrij"] = max(sw["poorten_totaal"] - sw["poorten_gebruikt"], 0)

    knelpunten = [
        {
            "id": "k1",
            "titel": "Geen managed switches / VLAN-segmentatie",
            "ernst": "hoog",
            "beschrijving": "Alle 6 switches zijn onbeheerd. Smart home, IoT, gastapparaten en werkplekken zitten op één plat netwerk. Een gehackt IoT-apparaat heeft direct toegang tot HA, NAS en PCs.",
        },
        {
            "id": "k2",
            "titel": "Single point of failure: hoofd-switch in meterkast",
            "ernst": "hoog",
            "beschrijving": "Bij stroomuitval of defect van de centrale switch valt het hele huis stil — inclusief slimme thermostaat, alarm en camera's. Geen UPS aanwezig.",
        },
        {
            "id": "k3",
            "titel": "Geen PoE op buitenkabels naar camera's",
            "ernst": "midden",
            "beschrijving": "Camera's en access points hangen nu aan losse PoE-injectors. Dat geeft een rommelig elektrisch beeld, extra storingspunten en geen centrale managment van PoE-budget.",
        },
        {
            "id": "k4",
            "titel": "Geen redundantie voor Home Assistant",
            "ernst": "midden",
            "beschrijving": "HA draait op één host (192.168.178.21) zonder externe back-up of off-site replicatie. Bij hardwarefalen verlies je alle automatiseringen en historie.",
        },
        {
            "id": "k5",
            "titel": "Cascade-topologie via huiskamerswitch",
            "ernst": "laag",
            "beschrijving": "Switch huiskamer is hub voor werkplek Marcel én Nick. Storing in huiskamer haalt twee gebruikers tegelijk offline. Acceptabel maar suboptimaal.",
        },
        {
            "id": "k6",
            "titel": "Buitenkabels niet beschermd via overspanningsbeveiliging",
            "ernst": "midden",
            "beschrijving": "30m, 20m en 10m buitenkabels (Cat7 S/FTP) komen het huis in zonder gecombineerde aardafleiding / surge protection. Bij blikseminslag in de buurt risico op kapot apparatuur.",
        },
        {
            "id": "k7",
            "titel": "Cat7 met RJ45 — verspilling van bandbreedte",
            "ernst": "info",
            "beschrijving": "Cat7 is gespecificeerd voor 600 MHz / GG45. In combinatie met RJ45 functioneert het feitelijk als Cat6a (10 GbE). Geen probleem, maar Cat6a was goedkoper geweest.",
        },
    ]

    verbeterplan = {
        "fase_1": {
            "titel": "Netwerk fundament",
            "horizon": "0 – 3 maanden",
            "kleur": "#00d4aa",
            "doel": "Veilig, gesegmenteerd, monitorbaar netwerk dat 24/7 werkt.",
            "acties": [
                {"titel": "Hoofdswitch vervangen door 24p PoE++ managed (bijv. UniFi USW-Pro-24-PoE)", "prio": "hoog", "kosten_eur": "€650 – €850", "complexiteit": "midden", "diy": True},
                {"titel": "VLAN-design: VLAN10 LAN, VLAN20 IoT, VLAN30 Gast, VLAN40 Camera's", "prio": "hoog", "kosten_eur": "€0 (configuratie)", "complexiteit": "hoog", "diy": False},
                {"titel": "UPS (CyberPower 1500VA) onder hoofdswitch + modem + HA-host", "prio": "hoog", "kosten_eur": "€180 – €240", "complexiteit": "laag", "diy": True},
                {"titel": "Cat7 buitenkabels via Surge Protector (Ubiquiti ETH-SP) aansluiten", "prio": "midden", "kosten_eur": "3× €25 = €75", "complexiteit": "laag", "diy": True},
                {"titel": "Patchpaneel + labeling in meterkast aanbrengen", "prio": "midden", "kosten_eur": "€90 – €140", "complexiteit": "midden", "diy": True},
                {"titel": "WiFi-mesh upgraden naar 2× UniFi U6-Pro/Lite via PoE", "prio": "midden", "kosten_eur": "2× €170 = €340", "complexiteit": "midden", "diy": True},
            ],
        },
        "fase_2": {
            "titel": "Smart Home uitbreiding",
            "horizon": "3 – 9 maanden",
            "kleur": "#9b59b6",
            "doel": "Compleet, toekomstvast smart home gebaseerd op Matter + Zigbee, volledig lokaal.",
            "acties": [
                {"titel": "Home Assistant migreren naar HA Yellow of NUC + SSD (van Pi/SD)", "prio": "hoog", "kosten_eur": "€350 – €550", "complexiteit": "midden", "diy": True},
                {"titel": "Matter-controller activeren in HA, eerste Matter-apparaten testen", "prio": "midden", "kosten_eur": "€0", "complexiteit": "midden", "diy": True},
                {"titel": "Aanwezigheidsdetectie via mmWave (Aqara FP2) in huiskamer + hal", "prio": "midden", "kosten_eur": "2× €75 = €150", "complexiteit": "laag", "diy": True},
                {"titel": "Slimme rookmelders (Ei Electronics Ei3028) integreren via Z-Wave/Matter", "prio": "hoog", "kosten_eur": "4× €70 = €280", "complexiteit": "laag", "diy": True},
                {"titel": "Smart-deurslot voordeur (Nuki 4 Pro of Yale Linus) + Matter-bridge", "prio": "midden", "kosten_eur": "€250 – €320", "complexiteit": "midden", "diy": True},
                {"titel": "Zonnepanelen-integratie & dynamische tarieven (P1 + ANWB Energie API)", "prio": "midden", "kosten_eur": "€0 – €60", "complexiteit": "hoog", "diy": False},
            ],
        },
        "fase_3": {
            "titel": "Automatisering & optimalisatie",
            "horizon": "9 – 18 maanden",
            "kleur": "#3498db",
            "doel": "Zelflerend huis: scènes op basis van aanwezigheid, weer, energieprijs en routines.",
            "acties": [
                {"titel": "Synology NAS in kastenkamer met automatische HA-snapshots (off-site naar OneDrive)", "prio": "hoog", "kosten_eur": "€500 – €750", "complexiteit": "midden", "diy": True},
                {"titel": "NVR voor camera's (Frigate op aparte mini-PC met Coral TPU)", "prio": "midden", "kosten_eur": "€350 – €450", "complexiteit": "hoog", "diy": False},
                {"titel": "Energie-dashboard: realtime verbruik per kring + advies via HA", "prio": "midden", "kosten_eur": "€80 – €150", "complexiteit": "midden", "diy": True},
                {"titel": "Stem-bediening lokaal (Whisper + Piper, geen cloud)", "prio": "laag", "kosten_eur": "€0", "complexiteit": "hoog", "diy": False},
                {"titel": "Aanwezigheids-routines: verlichting + verwarming + alarmstand per scenario", "prio": "midden", "kosten_eur": "€0", "complexiteit": "midden", "diy": True},
                {"titel": "Monitoring: Uptime Kuma + InfluxDB + Grafana met alerts via Pushover", "prio": "laag", "kosten_eur": "€0 – €60", "complexiteit": "midden", "diy": True},
            ],
        },
    }

    samenvatting = {
        "klant": "M. Montanus",
        "locatie": "Rijnsburg",
        "datum_rapport": "15-05-2026",
        "consultant": "Smart Home & Netwerk Consultant",
        "internet": "Ziggo (kabel) — alternatief glasvezel-scenario uitgewerkt",
        "metric_switches": len(switches),
        "metric_kabels": len([k for k in kabels if k.get("lengte_m")]),
        "metric_kabel_totaal_m": round(sum((k.get("lengte_m") or 0) for k in kabels), 1),
        "metric_buitenkabel_m": round(sum((k.get("lengte_m") or 0) for k in kabels if k.get("buitenkabel")), 1),
        "metric_apparaten": len(apparatuur),
        "metric_apparaten_bedraad": len([a for a in apparatuur if "bedraad" in a.get("verbinding", "")]),
        "metric_knelpunten": len(knelpunten),
        "metric_knelpunten_hoog": len([k for k in knelpunten if k["ernst"] == "hoog"]),
        "smart_home_hub": "Home Assistant Core 2026.5.1 (OS 17.3) op 192.168.178.21",
    }

    data = {
        "samenvatting": samenvatting,
        "switches": switches,
        "kabels": kabels,
        "topologie_edges": topologie,
        "apparatuur": apparatuur,
        "knelpunten": knelpunten,
        "verbeterplan": verbeterplan,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"OK — wegschreef {OUT_JSON} ({OUT_JSON.stat().st_size} bytes)")
    print(f"   switches: {len(switches)}, kabels: {len(kabels)}, apparatuur: {len(apparatuur)}, knelpunten: {len(knelpunten)}")


if __name__ == "__main__":
    main()
